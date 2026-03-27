import openpyxl
from collections import Counter

def process_excel_strict_v3(file_path, output_path):
    # 1. 加载 Excel
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Sheet1']

    # 2. 定位列索引
    header = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    try:
        idx_module = header.index("功能模块") + 1
        idx_point = header.index("功能点") + 1
    except ValueError:
        print(f"错误：未找到列名。当前表头：{header}")
        return

    # 3. 核心改进：统计【独立出现】的次数
    def count_distinct_occurrences(column_idx):
        distinct_names = []
        last_val = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=column_idx).value
            if val is not None and str(val).strip() != "":
                # 只有当发现一个新的值（即进入了新的合并块或新行）时，才记录
                distinct_names.append(str(val).strip())
        return Counter(distinct_names)

    # 分别获取模块和功能点的“独立出现”频次
    m_distinct_counts = count_distinct_occurrences(idx_module)
    p_distinct_counts = count_distinct_occurrences(idx_point)

    # 4. 第二次遍历：原地修改
    m_tracker = {}
    p_tracker = {}

    for row_idx in range(2, ws.max_row + 1):
        # --- 处理功能模块 ---
        m_cell = ws.cell(row=row_idx, column=idx_module)
        if m_cell.value is not None and str(m_cell.value).strip() != "":
            m_text = str(m_cell.value).strip()
            # 只有当这个名称在全表中【独立出现】超过 1 次，才编号
            if m_distinct_counts[m_text] > 1:
                m_tracker[m_text] = m_tracker.get(m_text, 0) + 1
                m_cell.value = f"{m_text}{m_tracker[m_text]}"

        # --- 处理功能点 ---
        p_cell = ws.cell(row=row_idx, column=idx_point)
        if p_cell.value is not None and str(p_cell.value).strip() != "":
            p_text = str(p_cell.value).strip()
            # 只有当这个名称在全表中【独立出现】超过 1 次，才编号
            if p_distinct_counts[p_text] > 1:
                p_tracker[p_text] = p_tracker.get(p_text, 0) + 1
                p_cell.value = f"{p_text}{p_tracker[p_text]}"

    # 5. 保存结果
    wb.save(output_path)
    print(f"处理完成！\n- 重复项已区分编号。\n- 唯一项（即便合并了多行）也不会带编号。\n文件存至：{output_path}")

# --- 路径配置 ---
input_xlsx = r"D:\DESKTOP\功能清单分工.xlsx"
output_xlsx = r"D:\DESKTOP\功能清单分工_已编号.xlsx"

if __name__ == "__main__":
    try:
        process_excel_strict_v3(input_xlsx, output_xlsx)
    except Exception as e:
        print(f"运行出错: {e}")